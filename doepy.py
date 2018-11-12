'''
Doepy (c) University of Manchester 2015

Doepy is licensed under the MIT License.

To view a copy of this license, visit <http://opensource.org/licenses/MIT/>.

@author:  Pablo Carbonell
@description: SYNBIOCHEM design of experiments
'''
from os import path, mkdir, system, unlink
import shutil, glob
import sys, re
import argparse
from datetime import datetime
import numpy as np
import pandas as pd
import sys
import csv
import json
import random
from viscad import viscad
import brOrligos
import rpy2
import rpy2.robjects as robjects

ID_COUNTER = 1

def construct(f):
    ct = []
    for l in open(f):
        m = l.rstrip().split('\t')
        if m[0].startswith('#'):
            continue
        factor = m[0]
        nlevels = int(m[1])
        try:
            positional = int(m[2])
        except:
            positional = 0
        try:
            pool = int(m[3])
        except:
            pool = 0
        # Degeneration: higher levels go to 1
        try:
            deg = int(m[4])
        except:
            deg = 0
        ct.append((factor, nlevels, positional, pool, deg))
    return ct


def convert_construct(xct, AddBlankPromoter=False):
    rid = {}
    ct = []
    for p in xct:
        comp = xct[p]['component']
        if comp == 'promoter':
            sbcid = False
            for l in xct[p]['levels']:
                if l.startswith('SBC'):
                    sbcid = True
                    break
            if sbcid:
                ll = []
                for l in xct[p]['levels']:
                    if l.startswith('SBC'):
                        lev = l
                    else:
                        lev = None
                    ll.append(lev)
                xct[p]['levels'] = ll
    multilevel = {}
    pairedFactor = {}
    for p in sorted(xct):
        levels = xct[p]['levels']
        comp = xct[p]['component']
        # deg is actual the number of possible values, 
        # in case of promoters we can have multiple empty values
        deg = 0
        if comp  == 'promoter':
            deg = len(levels) - len([i for i in filter( lambda h: h is None, levels)])
            if len(levels) > deg:
                deg += 1
        elif comp == 'gene':
            deg = len(levels) - len([i for i in filter( lambda h: h is None, levels)])
            if len(levels) ==  deg:
                deg = 0
        pos = xct[p]['positional']
        if pos is None:
            pos = 0
        cid = str(comp)+str(p)
        i = 1
        if AddBlankPromoter and \
           comp == 'promoter' and \
           len(levels) > len([i for i in filter( lambda h: h is None, levels)]):
            rid[cid+'_'+str(i)] = None
            i += 1
        if cid not in multilevel:
            multilevel[cid] = []
        if cid not in pairedFactor:
            pairedFactor[cid] = None
        for x in range(0, len(levels)):
            val = levels[x]
            valid = cid+'_'+str(i)
            # In case that some ids were already declared
            while valid in rid:
                i += 1
                valid = cid+'_'+str(i)
            rid[valid] = val
            # This is a little bit involved.
            # It should work as long as multigenes are added to higher secondary positions
            # as they are declared for the first time 
            try:
                multigene = xct[p]['multigene']
                if val in multigene:
                    for mg in multigene[val]:
                        mgpos, mgid = mg
                        mcid = str(comp)+str(mgpos)
                        valid = mcid+'_'+str(i)
                        rid[valid] = mgid
                        if mcid not in multilevel:
                            multilevel[mcid] = []
                        multilevel[mcid].append( valid )
                        pairedFactor[mcid] = cid
            except:
                pass
            i += 1
        ct.append((cid, len(levels)+len(multilevel[cid]), pos, 0, deg, pairedFactor[cid]))
    return ct, rid

# Transitional function to map old ids to new ICE ids
def map_oldid():
    import openpyxl
    midfile = '/mnt/SBC1/code/sbc-doe/mapping.xlsx'
    wb = openpyxl.load_workbook(midfile)
    xl = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    nl = xl.get_highest_row()
    mid = {}
    offset = 1
    for r in range(1, nl):
        newid = xl.cell(row=r, column= offset).value
        oldid = xl.cell(row=r, column= offset+1).value
        mid[oldid] = newid
    return mid


def write_fasta(fname, seqid, sequence):
    ow = open(fname, 'w')
    ow.write('>'+seqid+'\n')
    ow.write(sequence)
    ow.close()

def read_excel(e, s=1):
    df = pd.read_excel(e)
    mid = None
    seql = {}
    fact = {}
    partinfo = {}
    offset = 1
    fcol = 0
    r = 0
    multiGene = {}
    while fcol is not None:
        r += 1
        fcol = None
        try:
            fcol = df.iloc[r-1, offset-1]
            factor = int( fcol )
            positional = df.iloc[r-1, offset]
            component = str( df.iloc[r-1, offset+1] )
            part = str( df.iloc[r-1, offset+2] )
            partType =  str( df.iloc[r-1, offset+3] )
            if positional != positional:
                positional = None
        except:
            continue
        
        dependentFactor = False
        if component == 'gene':
            if len( partType.split(' ') ) > 1:
                if partType in multiGene:
                    # Multi-gene enzyme, add the information to the associated factor
                    dependentFactor = True
                    mfactor, mpart = multiGene[partType]
                    mgene = fact[mfactor]['multigene']
                    if mpart not in mgene:
                        mgene[mpart]= [ (factor, part) ]
                    else:
                        mgene[mpart].append( (factor, part) )
#                    continue
                else:
                    multiGene[partType] = (factor, part)

        # Legacy code, probably not useful
        if part is None:
            if factor in fact:
                i = len(fact[factor]['levels'])+1
            else:
                i = 1
            part = 'P'+str(factor)+'_'+str(i)
        seql[part] = None
        if part is not None and part.startswith('SBCPA'):
            if mid is None:
                mid = map_oldid()
            part = mid[part]
            seql[part] = None 
        if part == 'blank':
            part = None

        if factor not in fact:
            fact[factor] = {'positional': positional,
                            'component': component,
                            'levels': [],
                            'sequence': seql[part],
                            'multigene': {}
            }
        # add only as an independent level if it is not part of a previous multigene 
        if not dependentFactor:
            fact[factor]['levels'].append(part)
    return fact, seql, partinfo

def compact_factors(fact):
    """ This is a temporary solution for positional factors from read_json()
    At this point, we would accept only permutations for the positional factors.
    NOT COMPATIBLE with current definitions.
    It should be improved to give more flexibility.
    """
    positional = set()
    watch = set()
    for pos in fact:
        if fact[pos]['positional']:
            positional.add(pos)
            levels = fact[pos]['levels']
            fingerprint = ' '.join(sorted(fact[pos]['levels']))
            watch.add(fingerprint)
            if len(watch) > 1:
                raise Exception('Multiple positional factors') 
    if len(positional) != len(levels):
        raise Exception('More factors than slots')
    lpos = sorted(positional)
    llev = sorted(levels)
    for i in range(0, len(lpos)):
        fact[lpos[i]]['levels'] = [llev[i]]
    return fact
            

def read_json(f):
    """ Read json file return experimental design info.
    Initially the goal will be to replicate same result as with the Excel file.
    Not finished!!
    TO DO: templates are combinations that we want to keep! Modify code...,
    Positional factors are not correctly handled for current definitions
    """
    mid = None
    seql = {}
    fact = {}
    partinfo = {}
    jdata = json.load(open(f))
    collections = {}
    for col in jdata['plan']['collections']:
        collections[col['id']] = col
    constraints = jdata['plan']['specifications']['constraints']
    instances = {}
    for i in range(0, len(constraints)):
        for col in constraints[i]['collection']:
            if col not in instances:
                instances[col] = set()
            instances[col].add(i+1)
    for i in range(0, len(constraints)):
        factor = i + 1
        levels = []
        positional = 0
        for colId in constraints[i]['collection']:
            if len(instances[colId]) > 1:
                positional = 1
            item = collections[colId]
            # We assume no mix of types
            component = item['type']
            for partid in item['options']:
                part = partid.split('/')[-1]
                levels.append(part)
                partinfo[part] = {}
        if len(levels) > 0:
            fact[factor] = {'positional': positional,
                            'component': component,
                            'levels': levels,
                            'sequence': None
            }
    for part in partinfo:
        partinfo[part]['shortDescription'] = component
        partinfo[part]['name'] = part
    fact = compact_factors(fact)
    seed = int(jdata['seed'])
    return fact, partinfo, seed


def getfactors(ct, permut=False):
    factors =[]
    nlevels = []
    npos = []
    i = 0
    for x in ct:
        f = x[0]
        l = x[1]
        try:
            p = int(x[2])
        except:
            p = 0
        if l>1:
            factors.append(f)
            nlevels.append(l)
        if p != 0:
            npos.append(x[0])
        i += 1
    return factors, nlevels, npos

# Assume promoters "p"  and count number of segments
def segments(libr, ct):
    col = set()
    prom = set()
    for l in libr:
        x = ''
        pr = []
        for i in range(0, len(l)):
            ll = l[i]
            cti = ct[i][1]
            m = ll.split('_')
            if ll.startswith('p'):
                newsegment = False
                # If promoter has only 1 level, we assume always on
                if cti == 1:
                    newsegment = True
                    plevel = 2 # level 1 means off, level 2 on
                # else, level 1 means promoter off
                else:
                    # Promoter 1 is always on (add one to the level, 
                    # to be improved in a more general way)
                    if m[0] == 'p1':
                        newsegment = True
                        plevel = int(m[1]) + 1
                    elif m[1] != '1':
                        newsegment = True
                        plevel = int(m[1])
                if newsegment:
                    if x != '':
                        pr.append(x)
                        col.add(x)
#                    x = ''
                    x = str(plevel) # Put only the promoter level?
            elif ll.startswith('g'):
                x += ll
        col.add(x)
        pr.append(x)
        prom.add('.'.join(sorted(pr)))

    if x != '':
        col.add(x)

    sta= {}
    for i in col:
        m = i.split('_')
        l = len(m)
        if l not in sta:
            sta[l] = 0
        sta[l] += 1
    return col
#    return prom


def save_sbol(desid, libr, constructid, outfolder):
    print("Current implementation discontinued, to be updated with pysbol")
#    if not path.exists(outfolder):
#        mkdir(outfolder)
#    nid = []
#    x = sbol.sbol()
#    for i in range(0, len(libr)):
#        did = constructid[i]
#        x.construct3(did, libr[i])
#        out = path.join(outfolder, did+'.sbol')
#        x.serialize(out)
#    out = path.join(outfolder, desid+'.sbol')
#    x.serialize(out)
#    x = sbol.sbol()
#    x.collection(desid, constructid)
#    out = path.join(outfolder, desid+'_c'+'.sbol')
#    x.serialize(out)

# sbc id generator
def getsbcid(name, description, designid=None):
    """ Try to get design number if possible, otherwise keep the full label """
    """ Do not register in ICE, this is perfomed downstream in the pipeline """
    global ID_COUNTER
#    try:
#        desn = "DE%02d" % ( int( re.sub('^DE', '', re.sub('^.*SBC', '', designid) )), )
#    except:     
    desn = designid
                        
    partid = "%s_PL%02d" % (desn,ID_COUNTER)
    ID_COUNTER = ID_COUNTER + 1
    return partid



def save_design(design, ct, fname, lat, npos, rid = None, designid = None,
                constructid = [], partinfo = [], project=None, WriteCsv=None, WriteMap=None):
    ### Potentially the csv could be overwritten if multiple designs?
    ndes = {}
    n = 0
    # Read the design for each factor
    for x in ct:
        fact = x[0]
        found = False
        # Get the list of designed factors
        # For historical reasons (R), two types are possible
        if type(design['design']) == dict:
            flist = design['design'].keys()
        else:
            flist = design['design'].keys
                
        if fact in flist:
            ndes[fact] = np.array(design['design'][fact])
            n = len(ndes[fact])
        else:
            ndes[fact] = np.array([])
    # Note 01/18: Avoid this below, it is confusing and prone to errors.
    # It is better to use the information in the rid dictionary
    # coming from the DoE specification table
    # to know if we have an empty part (a promoter)
    # for x in ct:
    #     fact = x[0]
    #     nlevels = x[1]
    #     dege = x[4]
    #     if dege > 0:
    #         for i in range(0, len(ndes[fact])):
    #             if ndes[fact][i] > dege:
    #                 import pdb
    #                 pdb.set_trace()
    #                 ndes[fact][i] = 1

    
    for x in ct:
        fact = x[0]
        if len(ndes[fact]) == 0:
            ndes[fact] = np.repeat(1, n)
    # Add positional factor
    if 'pos' in flist:
        ndes['pos'] = np.array(design['design']['pos'])
    # Store designs
    of = open(fname, 'w')
    if WriteCsv is not None:
        cw = csv.writer(open(WriteCsv, 'w'), dialect='excel-tab' )
    if WriteMap is not None:
        mapw = csv.writer(open(WriteMap, 'w'))
    libr = []
    libscr = []
    for x in range(0, n):
        partsList = []
        screen = 1
        prom = False
        proml = []
        for counter in range(0, len(ct)):
            y = ct[counter]
            fa = y[0]
            le = y[1]
            po = y[2]
            pl = y[3]
            de = ndes[fa][x]
            # Screening pool?
            if pl > 0:
                if de > 1 or (le  == 1 and de > 0):
                    screen *= pl                        
            # Randomize permutations using a latin square
            # If only found at one position, ignore the permutation flag
            faid = "%s_%d" % (fa, ndes[fa][x],)
            if len(npos)> 1 and fa in npos:
                perm = ndes['pos'][x]
                fa = npos[lat[perm-1][npos.index(fa)]-1]
                faid = "%s_%d" % (fa, ndes[fa][x],)
            faid0 = faid
            if rid is not None:
                faid = rid[faid]

            if fa.startswith('promoter'):
                proml.append(counter)

            if faid is None:
                faid = ''
            else:
                # Keep track of last non-empty part to rule out two consecutive promoters
                if fa.startswith('promoter'):
                    if prom == True:
                        faid = ''
                    else:
                        prom = True
                else:
                    prom = False
            if faid is not None:
                partsList.append( ("%s" %  (faid,), faid0) )
                
        # Verify that we have a gene at the end
        for z in range(len(ct)-1,0,-1):
            if partsList[z][0] != '' and z not in proml:
                break
            elif z in proml:
                partsList[z] = ('', partsList[z][1])
        
        # Get the id
        if len(constructid) < x+1 :
            # Generate a meaningful name
            name = ''
            for partEntry in partsList:
                part = partEntry[0]
                if part != '':
                    if part in partinfo:
                        if name != '':
                            name += '-'
                        if partinfo[part]['shortDescription'].lower() == 'promoter':
                            name += '('
                        name += partinfo[part]['name']
                        if partinfo[part]['shortDescription'].lower() == 'promoter':
                            name += ')'

            if name == '':
                name = designid +'_'+str(x+1)
            
            description = 'Plasmid '
            if project is not None:
                description += project+'; '
            description += 'Design: '+ designid+'; '
            description += 'Construct: '+' '.join([i[0] for i in partsList])
            constructid.append(getsbcid(name, description, designid=designid))
        # Save the construct
        if rid is None:
            of.write("%s\t" % (constructid[x],))
            for part in partsList:
                of.write("%s\t" % (part[0],))
        else:
            if WriteCsv is not None:
                xx = []
                xx.append(constructid[x])
                for part in partsList:
                    if len(part[0]) > 0:
                        xx.append(part[0])
                cw.writerow(xx)
            if WriteMap is not None:
                xx = []
                xx.append(constructid[x])
                for part in partsList:
                    xx.append(':'.join(part))
                mapw.writerow(xx)
            of.write("%16s" % (constructid[x],))
            for part in partsList:
                of.write("%16s" % (part[0],))
        of.write('\n')
        libr.append([i[0] for i in partsList])
        if screen > 1:
            screen *= 3 # if screening a pool, multiply by 3
        libscr.append(screen)
    of.close()
    return libr, libscr

def save_seqs(outpath, constructid, libr, seql):
    for c in range(0, len(constructid)):
        seq = ''
        for s in libr[c]:
            if s != '':
                seq += seql[s]
        write_fasta(path.join(outpath, constructid[c]+'.fasta'), constructid[c], seq)

# If firstcolumn, the first column is the contruct name
def pcad(f, rid=None, firstcolumn=True, label=True,
         predefined='predefined_colors.txt', clean=True, nolabel=False):
    pcolors = {}
    if predefined is not None:
        for l in open(predefined):
            m = l.rstrip().split()
            pcolors[m[0]] = int(m[1])
    i = 0
    gl = []
    gl1 = []
    # Count how many levels each factor has
    count = {}
    for l in open(f):
        m = l.rstrip().split('\t')
        if firstcolumn:
            m = m[1:]
        for x in m:
            v = x.split('_')
            if v[0] not in gl:
                gl.append(v[0])
            if v[0] not in count:
                count[v[0]] = set()
            count[v[0]].add(x)
            if x not in gl1:
                gl1.append(x)
    if nolabel:
        nl = 'nl'
    else:
        nl = ' '
    fl = []
    labels = []
    for l in open(f):
        m = l.rstrip().split('\t')
        if firstcolumn:
            labels.append(m[0])
            m = m[1:]
        fn = f+'.pcad'+str(i)
        fl.append(fn)
        ow = open(fn, 'w')
        m = l.split()
        if firstcolumn:
            m = m[1:]
        for x in m:
            v = x.split('_')
            if x.startswith('promoter'):
                # p1: assumes promoter absence
                if v[1] != '1':
                    ow.write('t\n')
                    # For promoters, we just give promoter number and color it accordingly
                    if rid is not None and x in rid:
                        ow.write('p %s %d %s\n' % (rid[x],int(v[1])*2+2, nl))
                    else:
                        ow.write('p p%s %d %s\n' % (v[1],int(v[1])*2+2,nl))

            elif x.startswith('plasmid'):
    #                ow.write('p %s %d\n' % (v[0],gl.index(v[0])+1))
    # For promoters, we just give promoter number and color it accordingly
                if rid is not None and x in rid:
                    ow.write('p %s %d %s\n' % (rid[x],int(v[1])*2+2, nl))
                else:
                    ow.write('p p%s %d %s\n' % (v[1],int(v[1])*2+2, nl))
            else:
    #            ow.write('c %s %d\n' % (v[0],gl.index(v[0])+1))
                try:
                    color = gl1.index(x)+1
                except:
                    import pdb
                    pdb.set_trace()
                if rid is not None and x in rid:
                    name = rid[x]
                elif len(count[v[0]]) > 1:
                    name = x
                else:
                    name = v[0]
                if name in pcolors:
                    color = pcolors[name]
                ow.write('c %s %d %s\n' % (name,color,nl))
        ow.write('t\n')
        ow.write('# Arcs\n')
        ow.close()
        i += 1
    ofl = []
    ofs = []
    for i in range(0,len(fl)):
        pcad = fl[i]
        if label:
            ofl.append("label:'"+labels[i]+"'")
        of = pcad+'.png'
        ofl.append(of)
        ofsvg = pcad+'.svg'
        ofs.append(ofsvg)
        cmd = 'perl '+path.join(path.dirname(path.realpath(__file__)), 'piget.pl')+' '+pcad+' '+of
        print(cmd)
        system(cmd)

    cmd = 'convert '+' '.join(ofl)+' -append -flatten '+f+'.png'
    system(cmd)
    if clean:
        for x in fl+ofl+ofs:
            if path.exists(x):
                unlink(x)

def readJMP(jmp):
    header = None
    design = []
    doejmp = {'design': {}}
    for row in csv.reader(open(jmp, 'rU')):
        if header is None:
            header = row
            continue
        for i in range(0, len(header)):
            if len(row[i]) == 0:
                continue
            fact = header[i]
            # Skip for full factorial
            if re.search('Pattern',fact):
                continue
            if fact not in doejmp['design']:
                doejmp['design'][fact] = []
            try:
                val = int(row[i])
            except:
                val = int( re.sub('^L', '', row[i]) )
            doejmp['design'][fact].append( val )
    design.append(doejmp)
    return design

# def cleanDesign( doej, ct ):
#     """ For cases with empty genes, we can have cases where there are unnecessary promoters"""
#     """ Reduce to a single promoter """
#     newct = []
#     import pdb
#     pdb.set_trace()
#     for i in range(0, len(ct)):
#         f = ct[i]
#         # Current factor
#         factor = f[0]
#         import pdb
#         pdb.set_trace()
#         if factor.startswith('promoter'):
#             # No consecutive promoters
#             if prom:
#                 continue
#             prom = True
#             # No promoter at the end
#             if len(ct) == i+1:
#                 break
#         else:
#             prom = False
#         newct.append( f )
#     return newct
        
        

def addMultiGenes( doej, ct, rid ):
    """ Update the design in order to take into account multigenes.
        The information is updated based on ct """
    newct = []
    for f in ct:
        # Current factor
        factor = f[0]
        # Paired factor
        pair = f[5]
        # Number of levels
        nlev = f[1]
        addEmpty = False
        # Currently only work if this column is all dependent and therefore the design was empty
        if pair is not None:
            if factor in doej['design']:
                raise "Mixing multigenes and independent gene pairs not yet implemented"
            else:
                # Copy the paired factor, but if there is no available level, add an empty level
                doej['design'][factor] = []
                for x in doej['design'][pair]:
                    if x <= nlev:
                        doej['design'][factor].append( x )
                    else:
                        addEmpty = True
                        ncomp = factor+ '_'+ str(nlev+1)
                        rid[ncomp] = None
                        doej['design'][factor].append( nlev+1 )
                        
        # Redefine the factor with an additional empty level
        if addEmpty:
            f = list(f)
            f[1] += 1
            f = tuple(f)
                    
        newct.append( f )
    return doej, newct, rid

def doeconv(des):
    design = des[0]
#    resolution = des[1]
    des = { 'design': {} }
    factors = [n for n in design.names]
    for i in range(0, len(factors)):
        des['design'][factors[i]] = [i for i in design[i]]
    return des
        

def arguments():
    parser = argparse.ArgumentParser(description='SBC-DeO. Pablo Carbonell, SYNBIOCHEM, 2016')
    parser.add_argument('-p', action='store_true',
                        help='Full positional permutation (default: random latin square)')
    parser.add_argument('f', 
                        help='Input file with specifications (excel or txt format)')
    parser.add_argument('-s', default=1,
                        help='Excel sheet number (default 1)')
    parser.add_argument('-i', action='store_true',
                        help='Ignore segment calculations based on promoters')
    parser.add_argument('-r', action='store_false',
                        help='No regular fractional factorial design')
    parser.add_argument('-o', action='store_false',
                        help='No orthogonal array design')
    parser.add_argument('-x', nargs='?', type=int, default=100,
                        help='Random seed (default 100) [or pick random number] for oa design')
    parser.add_argument('id', 
                        help='Design id')
    parser.add_argument('-O',  
                        help='Output path')
    parser.add_argument('-b', action='store_false',  
                        help='Do not generate sbol file')
    parser.add_argument('-g', action='store_true',  
                        help='Generate pigeon cad image')
    parser.add_argument('-V', action='store_true',
                        help='Generate viscad diagram')
    parser.add_argument('-c', action='store_true',  
                        help='Generate construct fasta files')
    parser.add_argument('-v', 
                        help='Project description')
    parser.add_argument('-j', 
                        help='DoE from JMP')
    parser.add_argument('-w', action='store_true',
                        help='DoE from json (web version)')
    parser.add_argument('-G', 
                        help='Regenerate pigeon from file and exit')
    parser.add_argument('-k', action='store_false',
                        help='Keep pigeon files')
    parser.add_argument('-nolab', action='store_true',
                        help='Do not use labels in pigeon figures')
    parser.add_argument('-blankPromoter', action='store_true',
                        help='Add blank promoter even if not explicitly given')
    parser.add_argument('-bro', action='store_true',
                        help='Add file with full list of bridging oligos')
    return parser

def command_line(parser, args=None):
    if args is None:
        arg = parser.parse_args()
    else:
        arg = parser.parse_args(args)        
    return arg

def write_log(logfile, arg):
    s = []
    for x in arg:
        if len(x.split(' ')) > 1:
            s.append( '"{}"'.format(x) )
        else:
            s.append( x )
    with open(logfile, 'a') as handler:
        handler.write( ' '.join(s)+'\n' )

def run_doe(args=None):
    parser = arguments()
    arg = command_line(parser, args)
    f = arg.f
    p = arg.p
    cfasta = arg.c
    desid = arg.id
    outpath = arg.O
    sbolgen = arg.b
    cad = arg.g
    vcad = arg.V
    project = arg.v
    xarg = arg.x
    if outpath is None or not path.exists(outpath):
        outpath = path.dirname(f)
    outfolder = path.join(outpath)
    if not path.exists(outfolder):
        mkdir(outfolder)
    logfile = path.join(outfolder, desid+'.log')
    write_log(logfile, sys.argv)
    if arg.G is not None:
        rid = {}
        aa = []
        bb = []
        f1 = arg.G+'di0'
        f2 = arg.G+'d0'
        if path.exists(f1):
            for l in open(f1):
               m = l.rstrip().split('\t')
               for k in m:
                   bb.append(k)
        if path.exists(f2):
            for l in open(f2):
               ll = l.rstrip()
               for k in range(0, len(ll), 16):
                   val = l[k:(k+16)]
                   aa.append(re.sub(' ', '', val))
        for i in range(0, len(aa)):
            rid[bb[i]] = aa[i]
        pcad(f1, rid, clean=arg.k, nolabel=arg.nolab)
        sys.exit()
    if xarg is None:
        seed = np.random.randint(1e6)
    else:
        seed = xarg
    inputfile = path.join(outfolder, path.basename(f))
    if args is None:
        sys.argv[1] = '"'+path.basename(inputfile)+'"'
        cmd = ' '.join(['"{}"'.format(x) for x in sys.argv])
    else:
        cmd = ' '.join(['"{}"'.format(x) for x in args])
    s = int(arg.s)
    if not arg.w:
        try:
            xct, seql, partinfo = read_excel(inputfile, s)            
            ct, rid = convert_construct(xct, AddBlankPromoter=arg.blankPromoter)
        except:
            # old txt format (needs update)
            ct, cid = construct(inputfile)
            seql = {}
    else:
        xct, partinfo, seed = read_json(inputfile)
        ct, rid = convert_construct(xct, AddBlankPromoter=arg.blankPromoter)
    if arg.c:
        for s in seql:
            write_fasta(path.join(outpath, outfolder, s+'.fasta'), s, seql[s])
    wd = path.dirname(path.realpath(__file__))
    r = robjects.r
    r.source('mydeo.r')
    permut = r.permut
    rdoe1 = r.doe1
    rdoe2 = r.doe2
    # we keep only factors with more than one level
    # npos are the factors that can be rearranged
    factors, nlevels, npos = getfactors(ct)
    lat = None
    if len(npos) > 0:
        factors.append('pos')
        if not p:
            lat = np.array( permut(len(npos), ptype='latin') )
        else:
            lat = np.array( permut(len(npos), ptype='full') ) 
        lat = lat.astype(int)
        # add the levels corresponding to the shuffling
        nlevels.append(len(lat))
    if not p:
        designid = path.join(outfolder, desid)
    else:
        designid = path.join(outfolder, desid+'.full')
    constructid = []
    finfow = open(designid+'.info', 'w')
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    finfow.write('SBC-DoE; '+now+'\n')
    finfow.write(' Command: '+cmd+'\n')
    if not p:
        dinfo =  " Factors: %d; Levels: %d; Positional: %d [Latin square]" % (len(factors), np.prod(nlevels), len(npos))
    else:
        dinfo = " Factors: %d; Levels: %d; Positional: %d [Full permutations]" % (len(factors), np.prod(nlevels), len(npos))
    print('SBC-DoE; '+dinfo)
    finfow.write(dinfo+'\n')
    if arg.j is not None: # Custom design (column separated) or Factorial design using JMP
        if not path.exists(arg.j):
            arg.j = path.join(outfolder, arg.j)
        if not path.exists(arg.j):
            raise Exception('DoE file not found')
        jmp = arg.j
        doeJMP = readJMP(jmp)
        for i in range(0, len(doeJMP)):
            doeJMP[i], ct, rid = addMultiGenes( doeJMP[i], ct, rid )
#            doeJMP[i] = cleanDesign( doeJMP[i], ct )
        # This is a legacy loop, there should be a single JMP design
        for des in range(0, len(doeJMP)):
            if rid is not None:
                fname0 = designid+'.ji'+str(des)
                csvname = re.sub('\.[^.]*$', '.txt', fname0)
                fmapname = re.sub('\.[^.]*$', '.jmap', fname0)
                libr, libscr = save_design(doeJMP[des], ct, fname0, lat, npos, rid=rid, designid=desid, constructid=constructid, WriteCsv=csvname, WriteMap=fmapname)
            fname = designid+'.j'+str(des)
            libr, libscr = save_design(doeJMP[des],ct, fname, lat, npos, rid=None, designid=desid, constructid=constructid)
            if vars(arg)['i']:
                dinfor =  " Custom design %d; Library size: %d" % (des, len(libr))
            else:
                dinfor = " Custom design %d; Library size: %d; Segments: %d; Screening size: %d" % (des, len(libr), len(segments(libr, ct)), np.sum(libscr))
            print(dinfor)
            finfow.write(dinfor+'\n')
            if cad:
                pcad(fname, rid, clean=arg.k, nolabel=arg.nolab)
            if vcad:
#                viscad.runViscad(args=[fname, '-i', fname0, '-l', logfile])
                viscad.runViscad(args=[fmapname, '-l', logfile])
    if arg.r: # Regular fractional factorial design using R
        # Trivial case, no need of calling planor package
        if len(factors) == 1:
            doe1 = [{'design': {factors[0]: range(1, nlevels[0]+1) } } ]
        else:
            doe1rpy = rdoe1(factors=robjects.vectors.StrVector(factors), nlevels=nlevels, timeout=30)
            doe1 = []
            for des in doe1rpy:
                doe1.append( doeconv(des) )
        for des in range(0, len(doe1)):
            fnamer = designid+'.d'+str(des)
            libr, libscr = save_design(doe1[des], ct, fnamer, lat, npos,
                                       rid, desid, constructid, partinfo, project)
            if rid is not None:
                fnamer = designid+'.di'+str(des)
                libr, libscr = save_design(doe1[des], ct, fnamer, lat, npos,
                                           rid=None, designid=desid, constructid=constructid)
            if vars(arg)['i']:
                dinfor =  " Design %d; Model S^%d; Library size: %d" % (des, des+1, len(libr))
            else:
                dinfor = " Design %d; Model S^%d; Library size: %d; Segments: %d; Screening size: %d" % (des, des+1, len(libr), len(segments(libr, ct)), np.sum(libscr))
            print(dinfor)
            finfow.write(dinfor+'\n')
            if cad:
                pcad(fnamer, rid, clean=arg.k, nolabel=arg.nolab)
            if vcad:
                viscad.runViscad(args=[fnamer, '-i', fnamer, '-l', logfile, '-x', '_fr'])
    if arg.o: # Orthogonal arrays
        doe2rpy = rdoe2(factors=robjects.vectors.StrVector(factors), nlevels=robjects.vectors.IntVector(nlevels),
                           timeout=30, seed=seed)
        doe2 = []
        for des in doe2rpy:
            doe2.append( doeconv(des) )
        
        for des in range(0, len(doe2)):
            fnameo = designid+'.oad'+str(des)
            libr, libscr = save_design(doe2[des], ct, fnameo, lat, npos, rid, desid, constructid, partinfo, project)
            if cfasta:
                save_seqs(outfolder, constructid, libr, seql)
            if sbolgen:
                save_sbol(desid, libr, constructid, path.join(outfolder))
            if rid is not None:
                fnameo = designid+'.oadi'+str(des)
                libr, libscr = save_design(doe2[des], ct, fnameo, lat, npos, rid=None, designid=desid, constructid=constructid)
            if vars(arg)['i']:
                dinfor = " Orthogonal Array Design; Library size: %d; Seed: %d" % (len(libscr),seed)
            else:
                dinfor = " Orthogonal Array Design; Library size: %d; Segments: %d; Screening size: %d; Seed: %d" % (len(libr), len(segments(libr, ct)), np.sum(libscr), seed)
            print(dinfor)
            finfow.write(dinfor+'\n')
            if cad:
                pcad(fnameo, rid, clean=arg.k, nolabel=arg.nolab)
            if vcad:
                # Not working for OAs
                viscad.runViscad(args=[fnameo, '-i', fnameo, '-l', logfile, '-x', '_oa'])
    finfow.close()
    if arg.bro:
        # Needs some improvemet: not valid for all cases
        # Currently only implemented for custom designs (JMP)
        try:
            broFile = designid+'.bro'
            try:
                # improved version using the fmapname
                brArgs = [fmapname, fname, '-outFile', broFile, '-logFile', broFile+'.log']
            except:
                brArgs = [csvname, fname, '-outFile', broFile, '-logFile', broFile+'.log']
            brOrligos.run_bro( brArgs )
            write_log(logfile, ['broligos'] + brArgs)
        except:
            pass
    return outfolder, fname


if __name__ == '__main__':
    run_doe()
    
