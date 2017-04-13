'''
Doe (c) University of Manchester 2015

Doe is licensed under the MIT License.

To view a copy of this license, visit <http://opensource.org/licenses/MIT/>.

@author:  Pablo Carbonell
@description: SYNBIOCHEM design of experiments
'''
from os import path, mkdir, system, unlink
import shutil, glob
import sys, re
import argparse
from datetime import datetime
import pyRserve
import numpy as np
import doeopt
import sys
sys.path.append('/mnt/SBC1/code/sbc-api')   
import sbolutil as sbol
import sbcid
import iceutils
import csv
import json

def construct(f):
    ct = []
    for l in open(f):
        m = l.rstrip().split('\t')
        if m[0].startswith('#'):
            continue
        factor = m[0]
        nlevels = int(m[1])
        try:
            positional = int(m[2])
        except:
            positional = 0
        try:
            pool = int(m[3])
        except:
            pool = 0
        # Degeneration: higher levels go to 1
        try:
            deg = int(m[4])
        except:
            deg = 0
        ct.append((factor, nlevels, positional, pool, deg))
    return ct


def convert_construct(xct):
    rid = {}
    ct = []
    for p in xct:
        comp = xct[p]['component']
        if comp == 'promoter':
            sbcid = False
            for l in xct[p]['levels']:
                if l.startswith('SBC'):
                    sbcid = True
                    break
            if sbcid:
                ll = []
                for l in xct[p]['levels']:
                    if l.startswith('SBC'):
                        lev = l
                    else:
                        lev = None
                    ll.append(lev)
                xct[p]['levels'] = ll
    for p in sorted(xct):
        levels = xct[p]['levels']
        comp = xct[p]['component']
        # This is an exception for promoters, need to be improved
        deg = 0
        if comp  == 'promoter':
            deg = len(levels) - len(filter( lambda h: h is None, levels))
            if len(levels) > deg:
                deg += 1
        elif comp == 'gene':
            deg = len(levels) - len(filter( lambda h: h is None, levels))
            if len(levels) ==  deg:
                deg = 0
        pos = xct[p]['positional']
        if pos is None:
            pos = 0
        cid = str(xct[p]['component'])+str(p)
        i = 1
        if comp == 'promoter' and len(levels) > len(filter( lambda h: h is None, levels)):
            rid[cid+'_'+str(i)] = None
            i += 1
        for x in range(0, len(levels)):
            if levels[x] is not None:
                rid[cid+'_'+str(i)] = levels[x]
                i += 1
        ct.append((cid, len(levels), str(pos), 0, deg))
        
    return ct, rid

def get_sequence(partnumber):
    partid = int(re.sub('SBC', '', partnumber))
    ice = iceutils.ICESession(doeopt.ICE_RETRIEVE)
    return ice.get_sequence(partid)['sequence']

def get_part(partnumber):
    partid = int(re.sub('SBC', '', partnumber))
    ice = iceutils.ICESession(doeopt.ICE_RETRIEVE)
    return ice.get_part(partid)


# transitional function to map old ids to new ICE ids
def map_oldid():
    import openpyxl
    midfile = '/mnt/SBC1/code/sbc-doe/mapping.xlsx'
    wb = openpyxl.load_workbook(midfile)
    xl = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    nl = xl.get_highest_row()
    mid = {}
    for r in range(1, nl):
        newid = xl.cell(row=r, column= 0).value
        oldid = xl.cell(row=r, column= 1).value
        mid[oldid] = newid
    return mid


def write_fasta(fname, seqid, sequence):
    ow = open(fname, 'w')
    ow.write('>'+seqid+'\n')
    ow.write(sequence)
    ow.close()

def read_excel(e, s=1):
    import openpyxl
    wb = openpyxl.load_workbook(e)
    xl = wb.get_sheet_by_name(wb.get_sheet_names()[s-1])
    nl = xl.get_highest_row()
    mid = None
    seql = {}
    fact = {}
    partinfo = {}
    for r in range(1, nl):
        try:
            factor = int(xl.cell(row=r, column= 0).value)
            positional = xl.cell(row=r, column= 1).value
            component = xl.cell(row=r, column= 2).value
            part = xl.cell(row=r, column= 3).value
        except:
            continue
        if part is None:
            if factor in fact:
                i = len(fact[factor]['levels'])+1
            else:
                i = 1
            part = 'P'+str(factor)+'_'+str(i)
        seql[part] = None
        if part is not None and part.startswith('SBCPA'):
            if mid is None:
                mid = map_oldid()
            part = mid[part]
            seql[part] = get_sequence(part)
            partinfo[part] = get_part(part)
        if factor not in fact:
            fact[factor] = {'positional': positional,
                            'component': component,
                            'levels': [],
                            'sequence': seql[part]
            }
        if part == 'blank':
            part = None
        fact[factor]['levels'].append(part)
    return fact, seql, partinfo

def compact_factors(fact):
    """ This is a temporary solution for positional factors.
    At this point, we would accept only pure permutations for the positional factors.
    It should be improved to give more flexibility.
    """
    positional = set()
    watch = set()
    for pos in fact:
        if fact[pos]['positional']:
            positional.add(pos)
            levels = fact[pos]['levels']
            fingerprint = ' '.join(sorted(fact[pos]['levels']))
            watch.add(fingerprint)
            if len(watch) > 1:
                raise Exception('Multiple positional factors') 
    if len(positional) != len(levels):
        raise Exception('More factors than slots')
    lpos = sorted(positional)
    llev = sorted(levels)
    for i in range(0, len(lpos)):
        fact[lpos[i]]['levels'] = [llev[i]]
    return fact
            

def read_json(f):
    """ Read json file return experimental design info.
    Initially the goal will be to replicate same result as with the Excel file.
    TO DO: templates are combinations that we want to keep!! Modify code...
    """
    mid = None
    seql = {}
    fact = {}
    partinfo = {}
    jdata = json.load(open(f))
    collections = {}
    for col in jdata['plan']['collections']:
        collections[col['id']] = col
    constraints = jdata['plan']['specifications']['constraints']
    instances = {}
    for i in range(0, len(constraints)):
        for col in constraints[i]['collection']:
            if col not in instances:
                instances[col] = set()
            instances[col].add(i+1)
    for i in range(0, len(constraints)):
        factor = i + 1
        levels = []
        positional = 0
        for colId in constraints[i]['collection']:
            if len(instances[colId]) > 1:
                positional = 1
            item = collections[colId]
            # We assume no mix of types
            component = item['type']
            for partid in item['options']:
                part = partid.split('/')[-1]
                levels.append(part)
                partinfo[part] = {}
        if len(levels) > 0:
            fact[factor] = {'positional': positional,
                            'component': component,
                            'levels': levels,
                            'sequence': None
            }
    for part in partinfo:
        partinfo[part]['shortDescription'] = component
        partinfo[part]['name'] = part
    fact = compact_factors(fact)
    seed = int(jdata['seed'])
    return fact, partinfo, seed


def getfactors(ct, permut=False):
    factors =[]
    nlevels = []
    npos = []
    i = 0
    for x in ct:
        f = x[0]
        l = x[1]
        try:
            p = int(x[2])
        except:
            p = 0
        if l>1:
            factors.append(f)
            nlevels.append(l)
        if p != 0:
            npos.append(x[0])
        i += 1
    return factors, nlevels, npos

# Assume promoters "p"  and count number of segments
def segments(libr, ct):
    col = set()
    prom = set()
    for l in libr:
        x = ''
        pr = []
        for i in range(0, len(l)):
            ll = l[i]
            cti = ct[i][1]
            m = ll.split('_')
            if ll.startswith('p'):
                newsegment = False
                # If promoter has only 1 level, we assume always on
                if cti == 1:
                    newsegment = True
                    plevel = 2 # level 1 means off, level 2 on
                # else, level 1 means promoter off
                else:
                    # Promoter 1 is always on (add one to the level, to be improved in a more general way)
                    if m[0] == 'p1':
                        newsegment = True
                        plevel = int(m[1]) + 1
                    elif m[1] != '1':
                        newsegment = True
                        plevel = int(m[1])
                if newsegment:
                    if x != '':
                        pr.append(x)
                        col.add(x)
#                    x = ''
                    x = str(plevel) # Put only the promoter level?
            elif ll.startswith('g'):
                x += ll
        col.add(x)
        pr.append(x)
        prom.add('.'.join(sorted(pr)))

    if x != '':
        col.add(x)

    sta= {}
    for i in col:
        m = i.split('_')
        l = len(m)
        if l not in sta:
            sta[l] = 0
        sta[l] += 1
    return col
#    return prom


def save_sbol(desid, libr, constructid, outfolder):
    if not path.exists(outfolder):
        mkdir(outfolder)
    nid = []
    x = sbol.sbol()
    for i in range(0, len(libr)):
        did = constructid[i]
        x.construct3(did, libr[i])
#        out = path.join(outfolder, did+'.sbol')
#        x.serialize(out)
    out = path.join(outfolder, desid+'.sbol')
    x.serialize(out)
#    x = sbol.sbol()
#    x.collection(desid, constructid)
#    out = path.join(outfolder, desid+'_c'+'.sbol')
#    x.serialize(out)

# sbc id generator
def getsbcid(name, description, RegisterinICE= False, designid=None):
    if RegisterinICE:
        responsible = doeopt.ICE_USER
        email = doeopt.ICE_EMAIL
        ice = iceutils.ICESession(doeopt.ICE_POST, info=doeopt.ICE_INFO) # icetest
        plasmid = ice.template_newplasmid(name , description, responsible, responsible, responsible, email, email, email)
        reg_plasmid = ice.create_part(plasmid)
        sbcID = reg_plasmid['id']
        group_number = 2
        ice.add_write_permission(sbcID, group_number)
        partid = ice.get_part(sbcID)['partId']
    else:
        response = sbcid.reserve('DE', 1, doeopt.ICE_EMAIL, 'Construct in combinatorial library '+designid)
        partid = "SBCDE%06d" % (response['result'][0]['id'],)
    return partid



def save_design(design, ct, fname, lat, npos, rid = None, designid = None, constructid = [], partinfo = [], project=None, RegisterinIce=False):
    ndes = {}
    n = 0
    # Read the design for each factor
    for x in ct:
        fact = x[0]
        found = False
        if type(design['design']) == dict:
            flist = design['design'].keys()
        else:
            flist = design['design'].keys
                
        if fact in flist:
            ndes[fact] = np.array(design['design'][fact])
            n = len(ndes[fact])
        else:
            ndes[fact] = np.array([])
    for x in ct:
        fact = x[0]
        nlevels = x[1]
        dege = x[4]
        if dege > 0:
            for i in range(0, len(ndes[fact])):
                if ndes[fact][i] > dege:
                    ndes[fact][i] = 1
    
    for x in ct:
        fact = x[0]
        if len(ndes[fact]) == 0:
            ndes[fact] = np.repeat(1, n)
    # Add positional factor
    if 'pos' in flist:
        ndes['pos'] = np.array(design['design']['pos'])
    # Store designs
    of = open(fname, 'w')
    libr = []
    libscr = []
    for x in range(0, n):
        ll = []
        screen = 1
        for y in ct:
            fa = y[0]
            le = y[1]
            po = y[2]
            pl = y[3]
            de = ndes[fa][x]
            # Screening pool?
            if pl > 0:
                if de > 1 or (le  == 1 and de > 0):
                    screen *= pl                        
            # Randomize permutations using a latin square
            faid = "%s_%d" % (fa, ndes[fa][x],)
            if fa in npos:
                perm = ndes['pos'][x]
                fa = npos[lat[perm-1][npos.index(fa)]-1]
                faid = "%s_%d" % (fa, ndes[fa][x],)
            if rid is not None:
                faid = rid[faid]
            if faid is None:
                faid = ''
            if faid is not None:
                ll.append("%s" %  (faid,))

        # Get the id
        if len(constructid) < x+1 :
            # Generate a meaningful name
            name = ''
            for part in ll:
                if part != '':
                    if part in partinfo:
                        if name != '':
                            name += '-'
                        if partinfo[part]['shortDescription'].lower() == 'promoter':
                            name += '('
                        name += partinfo[part]['name']
                        if partinfo[part]['shortDescription'].lower() == 'promoter':
                            name += ')'

            if name == '':
                name = designid +'_'+str(x+1)
            
            description = 'Plasmid '
            if project is not None:
                description += project+'; '
            description += 'Design: '+ designid+'; '
            description += 'Construct: '+' '.join(ll)
            constructid.append(getsbcid(name, description, RegisterinICE=RegisterinIce, designid=designid))
        # Save the construct
        if rid is None:
            of.write("%s\t" % (constructid[x],))
            for part in ll:
                of.write("%s\t" % (part,))
        else:
            of.write("%16s" % (constructid[x],))
            for part in ll:
                of.write("%16s" % (part,))
        of.write('\n')
        libr.append(ll)
        if screen > 1:
            screen *= 3 # if screening a pool, multiply by 3
        libscr.append(screen)
    of.close()
    return libr, libscr

def save_seqs(outpath, constructid, libr, seql):
    for c in range(0, len(constructid)):
        seq = ''
        for s in libr[c]:
            if s != '':
                seq += seql[s]
        write_fasta(path.join(outpath, constructid[c]+'.fasta'), constructid[c], seq)

# If firstcolumn, the first column is the contruct name
def pcad(f, rid=None, firstcolumn=True, label=True, predefined='predefined_colors.txt'):
    pcolors = {}
    if predefined is not None:
        for l in open(predefined):
            m = l.rstrip().split()
            pcolors[m[0]] = int(m[1])
    i = 0
    gl = []
    gl1 = []
    # Count how many levels each factor has
    count = {}
    for l in open(f):
        m = l.rstrip().split('\t')
        if firstcolumn:
            m = m[1:]
        for x in m:
            v = x.split('_')
            if v[0] not in gl:
                gl.append(v[0])
            if v[0] not in count:
                count[v[0]] = set()
            count[v[0]].add(x)
            if x not in gl1:
                gl1.append(x)
    fl = []
    labels = []
    for l in open(f):
        m = l.rstrip().split('\t')
        if firstcolumn:
            labels.append(m[0])
            m = m[1:]
        fn = f+'.pcad'+str(i)
        fl.append(fn)
        ow = open(fn, 'w')
        m = l.split()
        if firstcolumn:
            m = m[1:]
        for x in m:
            v = x.split('_')
            if x.startswith('promoter'):
                # p1: assumes promoter absence
                if v[1] != '1':
                    ow.write('t\n')
                    # For promoters, we just give promoter number and color it accordingly
                    if rid is not None and x in rid:
                        ow.write('p %s %d\n' % (rid[x],int(v[1])*2+2))
                    else:
                        ow.write('p p%s %d\n' % (v[1],int(v[1])*2+2))

            elif x.startswith('plasmid'):
    #                ow.write('p %s %d\n' % (v[0],gl.index(v[0])+1))
    # For promoters, we just give promoter number and color it accordingly
                if rid is not None and x in rid:
                    ow.write('p %s %d\n' % (rid[x],int(v[1])*2+2))
                else:
                    ow.write('p p%s %d\n' % (v[1],int(v[1])*2+2))
            else:
    #            ow.write('c %s %d\n' % (v[0],gl.index(v[0])+1))
                try:
                    color = gl1.index(x)+1
                except:
                    import pdb
                    pdb.set_trace()
                if rid is not None and x in rid:
                    name = rid[x]
                elif len(count[v[0]]) > 1:
                    name = x
                else:
                    name = v[0]
                if name in pcolors:
                    color = pcolors[name]
                ow.write('c %s %d\n' % (name,color))
        ow.write('t\n')
        ow.write('# Arcs\n')
        ow.close()
        i += 1

    ofl = []
    for i in range(0,len(fl)):
        pcad = fl[i]
        if label:
            ofl.append("label:'"+labels[i]+"'")
        of = pcad+'.png'
        ofl.append(of)
        cmd = 'perl '+path.join(path.dirname(path.realpath(__file__)), 'piget.pl')+' '+pcad+' '+of
        system(cmd)

    cmd = 'convert '+' '.join(ofl)+' -append '+f+'.png'
    system(cmd)
    for x in fl+ofl:
        if path.exists(x):
            unlink(x)

def readJMP(jmp):
    header = None
    design = []
    doejmp = {'design': {}}
    for row in csv.reader(open(jmp)):
        if header is None:
            header = row
            continue
        for i in range(0, len(header)):
            fact = header[i]
            if fact not in doejmp['design']:
                doejmp['design'][fact] = []
            doejmp['design'][fact].append(int(row[i]))
    design.append(doejmp)
    return design

def arguments():
    parser = argparse.ArgumentParser(description='SBC-DeO. Pablo Carbonell, SYNBIOCHEM, 2016')
    parser.add_argument('-p', action='store_true',
                        help='Full positional permutation (default: random latin square)')
    parser.add_argument('f', 
                        help='Input file with specifications (excel or txt format)')
    parser.add_argument('-s', default=1,
                        help='Excel sheet number (default 1)')
    parser.add_argument('-i', action='store_true',
                        help='Ignore segment calculations based on promoters')
    parser.add_argument('-r', action='store_false',
                        help='No regular fractional factorial design')
    parser.add_argument('-o', action='store_false',
                        help='No orthogonal array design')
    parser.add_argument('-x', nargs='?', type=int, default=100,
                        help='Random seed (default 100) [or pick random number] for oa design')
    parser.add_argument('id', 
                        help='Design id')
    parser.add_argument('-O',  
                        help='Output path')
    parser.add_argument('-b', action='store_false',  
                        help='Do not generate sbol file')
    parser.add_argument('-g', action='store_true',  
                        help='Generate pigeon cad image')
    parser.add_argument('-c', action='store_true',  
                        help='Generate construct fasta files')
    parser.add_argument('-v', 
                        help='Project description')
    parser.add_argument('-I', action='store_true',
                        help='Register project in ICE [False]')
    parser.add_argument('-j', 
                        help='DoE from JMP')
    parser.add_argument('-w', action='store_true',
                        help='DoE from json (web version)')
    parser.add_argument('-G', 
                        help='Regenerate pigeon from file and exit')
    return parser

def command_line(parser, args=None):
    if args is None:
        arg = parser.parse_args()
    else:
        arg = parser.parse_args(args)        
    return arg


def run_doe(args=None):
    parser = arguments()
    arg = command_line(parser, args)
    if arg.G is not None:
        rid = {}
        aa = []
        bb = []
        f1 = arg.G+'.di0'
        f2 = arg.G+'.d0'
        if path.exists(f1):
            for l in open(f1):
               m = l.rstrip().split('\t')
               for k in m:
                   bb.append(k)
        if path.exists(f2):
            for l in open(f2):
               ll = l.rstrip()
               for k in range(0, len(ll), 16):
                   val = l[k:(k+16)]
                   aa.append(re.sub(' ', '', val))
        for i in range(0, len(aa)):
            rid[bb[i]] = aa[i]
        pcad(f1, rid)
        sys.exit()
    f = arg.f
    p = arg.p
    cfasta = arg.c
    desid = arg.id
    outpath = arg.O
    sbolgen = arg.b
    cad = arg.g
    project = arg.v
    xarg = arg.x
    if xarg is None:
        seed = np.random.randint(1e6)
    else:
        seed = xarg
    if outpath is None or not path.exists(outpath):
        outpath = path.dirname(f)
    outfolder = path.join(outpath, desid)
    if not path.exists(outfolder):
        mkdir(outfolder)
    inputfile = path.join(outfolder, path.basename(f))
    shutil.copyfile(f, inputfile)
    if args is None:
        sys.argv[1] = '"'+path.basename(inputfile)+'"'
        cmd = ' '.join(sys.argv)
    else:
        cmd = ' '.join(args)
    s = int(arg.s)
    if not arg.w:
        try:
            xct, seql, partinfo = read_excel(inputfile, s)
            ct, rid = convert_construct(xct)
        except:
            # old txt format (needs update)
            ct, cid = construct(inputfile)
            seql = {}
    else:
        xct, partinfo, seed = read_json(inputfile)
        ct, rid = convert_construct(xct)
    if arg.c:
        for s in seql:
            write_fasta(path.join(outpath, outfolder, s+'.fasta'), s, seql[s])
    wd = path.dirname(path.realpath(__file__))
    conn = pyRserve.connect()
    conn.r.source(path.join(wd, 'mydeo.r'))
    factors, nlevels, npos = getfactors(ct)
    lat = None
    if len(npos) > 0:
        factors.append('pos')
        if not p:
            lat = conn.r.permut(len(npos), ptype='latin')
        else:
            lat = np.array(conn.r.permut(len(npos), ptype='full'))
        lat = lat.astype(int)
        nlevels.append(len(lat))
    if not p:
        designid = path.join(outfolder, desid)
    else:
        designid = path.join(outfolder, desid+'.full')
    constructid = []
    finfow = open(designid+'.info', 'w')
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    finfow.write('SBC-DoE; '+now+'\n')
    finfow.write(' Command: '+cmd+'\n')
    if not p:
        dinfo =  " Factors: %d; Levels: %d; Positional: %d [Latin square]" % (len(factors), np.prod(nlevels), len(npos))
    else:
        dinfo = " Factors: %d; Levels: %d; Positional: %d [Full permutations]" % (len(factors), np.prod(nlevels), len(npos))
    print('SBC-DoE; '+dinfo)
    finfow.write(dinfo+'\n')
    if arg.j is not None:
        jmp = arg.j
        doeJMP = readJMP(jmp)
        for des in range(0, len(doeJMP)):
            if rid is not None:
                fname = designid+'.ji'+str(des)
                libr, libscr = save_design(doeJMP[des], ct, fname, lat, npos, rid=rid, designid=desid, constructid=constructid, RegisterinIce=arg.I)
            fname = designid+'.j'+str(des)
            libr, libscr = save_design(doeJMP[des],ct, fname, lat, npos, rid=None, designid=desid, constructid=constructid, RegisterinIce=arg.I)
            if vars(arg)['i']:
                dinfor =  " Custom design %d; Library size: %d" % (des, len(libr))
            else:
                dinfor = " Custom design %d; Library size: %d; Segments: %d; Screening size: %d" % (des, len(libr), len(segments(libr, ct)), np.sum(libscr))
            print(dinfor)
            finfow.write(dinfor+'\n')
            if cad:
                pcad(fname, rid)
    if arg.r:
        doe1 = conn.r.doe1(factors=np.array(factors), nlevels=np.array(nlevels), timeout=30)
        for des in range(0, len(doe1)):
            fname = designid+'.d'+str(des)
            libr, libscr = save_design(doe1[des], ct, fname, lat, npos, rid, desid, constructid, partinfo, project, RegisterinIce=arg.I)
            if rid is not None:
                fname = designid+'.di'+str(des)
                libr, libscr = save_design(doe1[des], ct, fname, lat, npos, rid=None, designid=desid, constructid=constructid, RegisterinIce=arg.I)
            if vars(arg)['i']:
                dinfor =  " Design %d; Model S^%d; Library size: %d" % (des, des+1, len(libr))
            else:
                dinfor = " Design %d; Model S^%d; Library size: %d; Segments: %d; Screening size: %d" % (des, des+1, len(libr), len(segments(libr, ct)), np.sum(libscr))
            print(dinfor)
            finfow.write(dinfor+'\n')
            if cad:
                pcad(fname, rid)
    if arg.o:
        doe2 = conn.r.doe2(factors=np.array(factors), nlevels=np.array(nlevels),
                           timeout=30, seed=seed)

        for des in range(0, len(doe2)):
            fname = designid+'.oad'+str(des)
            libr, libscr = save_design(doe2[des], ct, fname, lat, npos, rid, desid, constructid, partinfo, project, RegisterinIce=arg.I)
            if cfasta:
                save_seqs(outfolder, constructid, libr, seql)
            if sbolgen:
                save_sbol(desid, libr, constructid, path.join(outfolder))
            if rid is not None:
                fname = designid+'.oadi'+str(des)
                libr, libscr = save_design(doe2[des], ct, fname, lat, npos, rid=None, designid=desid, constructid=constructid, RegisterinIce=arg.I)
            if vars(arg)['i']:
                dinfor = " Orthogonal Array Design; Library size: %d; Seed: %d" % (len(libscr),seed)
            else:
                dinfor = " Orthogonal Array Design; Library size: %d; Segments: %d; Screening size: %d; Seed: %d" % (len(libr), len(segments(libr, ct)), np.sum(libscr), seed)
            print(dinfor)
            finfow.write(dinfor+'\n')
            if cad:
                pcad(fname, rid)
    finfow.close()
    return outfolder, fname


if __name__ == '__main__':
    run_doe()
